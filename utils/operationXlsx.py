# -*- encoding:utf-8 -*-
"""
@作者：wuchengan
@文件名：operationXlsx.py
@时间：2019/5/23  14:02
@文档说明:
"""

import openpyxl,os,xlrd
from openpyxl.styles import Font,PatternFill
from openpyxl.styles.colors import Color,RED,GREEN,BLACK,BLUE,YELLOW,DARKGREEN

class OperationExcel(object):
    def dir_base(self, filename, filePath='data'):
        """
        获取data文件夹下的文件
        :param filename: 要读取的文件名称
        :param filePath: 要读取的文件对应的文件夹
        :return:
        """
        return os.path.join(os.path.dirname(os.path.dirname(__file__)), filePath, filename)
    
    def getExcelData(self, filename):
        """
        获取Excel数据
        :param filename:Excel文件名
        :return: rows
        """
        rows = []
        # wb=openpyxl.load_workbook(self.dir_base(filename))
        # sheet = wb.worksheets[0]  # 获取excel第一张sheet表
        # for row_idx in sheet.rows:
        #     row=[cell.value for cell in row_idx]
        #     rows.append(row)
        with xlrd.open_workbook(self.dir_base(filename)) as f:
            sheet=f.sheet_by_index(0) #获取excel第一张sheet表
            for row_idx in range(1,sheet.nrows):
                rows.append(sheet.row_values(row_idx,0,sheet.ncols-1))
            return rows
    
    def modifyExcel(self,filename,row,col,result,reason):
        """
        修改xlsx的单元格数据
        :param filename: 文件名
        :param local: 修改的单元格位置，例如A3,B2
        :param result: 单元格修改的值
        :return:
        """
        # 打开工作表
        work_book=openpyxl.load_workbook(self.dir_base(filename))
        # 获取工作表sheet页
        sheet=work_book.worksheets[0]
        # 给目标单元格赋值
        sheet.cell(row,col).value=result
        #增加备注，执行失败时将失败原因抛出在单元格
        sheet.cell(row, col + 1).value = reason
        if result == 'P':
            # 将结果加粗
            sheet.cell(row,col).font = Font(bold=True)
            # 将结果单元格填充颜色，solid意思是纯色填充，还有其他形式，自行查
            sheet.cell(row, col).fill = PatternFill('solid',fgColor=GREEN)
        if result == 'F':
            sheet.cell(row,col).font = Font(bold=True)
            sheet.cell(row, col).fill = PatternFill('solid', fgColor=RED)
        work_book.save(self.dir_base(filename))

# e=OperationExcel()
# v=e.getExcelData('data.xlsx')
# print(v)
